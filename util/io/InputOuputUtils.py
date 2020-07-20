# -*- coding: utf-8 -*-
# @File         : InputOutputUtils.py
# @Author       : Zhendong Zhang
# @Organization : Huazhong University of Science and Technology
# @Email        : zzd_zzd@hust.edu.cn
# @Time         : 2020/6/28 22:05 

import pandas as pd
from openpyxl import load_workbook
import os
import numpy as np


class JsonInputOutputUtils(object):

    @staticmethod
    def saveJson(jsonData, savePath):
        with open(savePath, 'w') as f:
            f.write(jsonData)
            f.close()


class NumpyInputOutputUtils(object):

    @staticmethod
    def saveNumpyToExcel(numpyData, savePath, sheetName):
        pandasData = pd.DataFrame(numpyData)
        PandasInputOutputUtils.savePandasToExcel(pandasData, savePath, sheetName)

    @staticmethod
    def readExcelDataToNumpy(excelPath, sheetName):
        pandasData = PandasInputOutputUtils.readExcelDataToPandas(excelPath, sheetName)
        numpyData = np.array(pandasData)
        return numpyData


class PandasInputOutputUtils(object):

    @staticmethod
    def savePandasToExcel(pandasData, savePath, sheetName):
        dirPath = os.path.dirname(savePath)
        if not os.path.exists(dirPath):
            os.makedirs(dirPath)
        # 这种写法不会覆盖之前的sheet
        writer = pd.ExcelWriter(savePath, engine='openpyxl')
        if os.path.exists(savePath):
            book = load_workbook(savePath)
            if sheetName in book.sheetnames:
                book.remove(book[sheetName])
            writer.book = book
        pandasData.to_excel(writer, sheetName)
        writer.save()
        writer.close()

    @staticmethod
    def readExcelDataToPandas(excelPath, sheetName):
        pandasData = pd.read_excel(excelPath, sheet_name=sheetName)
        return pandasData
